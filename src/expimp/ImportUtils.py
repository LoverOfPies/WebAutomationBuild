import json

from openpyxl import load_workbook

from app import db
from src.utils import get_file_name, get_model, get_condition


def get_or_insert(model, values):
    condition = get_condition(model, values)
    if condition is None:
        return None
    try:
        obj = model.get(condition)
    except model.DoesNotExist:
        with db.database.atomic() as transaction:
            try:
                obj = model.insert(values).execute()
                transaction.commit()
            except:
                transaction.rollback()
                return None
    return obj


def import_single_table(model, file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    fields = {}
    data = []
    for row in sheet:
        value = {}
        for cell in row:
            if cell.row == 1:
                if not hasattr(model, cell.value):
                    return False
                fields[cell.column] = cell.value
                continue
            value[fields[cell.column]] = cell.value
        if len(value) > 0:
            data.append(value)
    with db.database.atomic() as transaction:
        try:
            model.insert(data).execute()
            transaction.commit()
        except:
            transaction.rollback()
            return False
    return True


def import_custom_data(file_path):
    file_name = get_file_name(file_path)
    with open('src/expimp/custom_config.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    config = data[file_name]
    if config is None:
        return False
    if file_name.lower() == 'Материалы'.lower():
        import_material_table(file_path, config)
        return True
    return False


def import_material_table(file_path, config):
    wb = load_workbook(file_path)
    sheet = wb.active
    for row in sheet:
        value_material_category = None
        value_material_group = None
        value_material_subgroup = None
        material = None
        value = {}
        for cell in row:
            if cell.row == 1:
                continue
            table_name, field = config[str(cell.column)].split(".")
            value[field] = cell.value
            if cell.column == 2:
                model = get_model(table_name)
                value_material_category = get_or_insert(model, value)
                value = {}
            if cell.column == 4:
                model = get_model(table_name)
                value["material_category"] = value_material_category
                value_material_group = get_or_insert(model, value)
                value = {}
            if cell.column == 6:
                model = get_model(table_name)
                value["material_group"] = value_material_group
                value_material_subgroup = get_or_insert(model, value)
                value = {}
            if cell.column == 10:
                model = get_model(table_name)
                value["material_subgroup"] = value_material_subgroup
                unit = get_or_insert(get_model("unit"), {"name": value["unit"]})
                value["unit"] = unit
                material = get_or_insert(model, value)
                value = {}
