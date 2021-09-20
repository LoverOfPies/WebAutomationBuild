from openpyxl import Workbook

from src.db.models.BaseModel import BaseModel


def export_table(model):
    model_meta = model._meta
    fields = [key for key in model_meta.fields]
    fields.remove('id')
    data = model.select()
    table_name = model_meta.name
    wb = Workbook()
    ws = wb.active
    ws.title = table_name
    for col in range(len(fields)):
        ws.cell(row=1, column=col + 1, value=fields[col])
    row = 2
    for attribute in data:
        for col in range(len(fields)):
            val = getattr(attribute, fields[col])
            if isinstance(val, BaseModel):
                val = getattr(val, 'uuid')
            ws.cell(row=row, column=col + 1, value=str(val))
        row += 1
    wb.save(f'expimp/{table_name}.xlsx')
