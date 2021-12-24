import os

from openpyxl import load_workbook, Workbook
from peewee import ForeignKeyField, Model
from werkzeug.security import safe_join
from werkzeug.utils import secure_filename

from app import app
from src.db import DataBaseUtils


def allowed_file(file_ext, extensions):
    return file_ext in app.config[extensions]


def get_path(filename):
    return safe_join(app.config['UPLOAD_FOLDER'], filename)


def get_file_name(path):
    f_name, _ = os.path.splitext(os.path.basename(path))
    return f_name


def get_file_ext(path):
    _, f_ext = os.path.splitext(os.path.basename(path))
    return f_ext


def save_file(files):
    if 'file' not in files:
        return None
    file = files['file']
    if file.filename == '':
        return None
    if file and allowed_file(get_file_ext(file.filename), 'EXCEL_EXTENSIONS'):
        filename = secure_filename(file.filename)
        path = get_path(filename)
        file.save(path)
        return path
    return None


def create_file(collection):
    path = collection
    return path


def import_single_table(collection):
    model = DataBaseUtils.get_model(collection)
    wb = load_workbook(f'expimp/{collection}.xlsx')
    sheet = wb.active
    fields = {}
    data = []
    for row in sheet:
        value = {}
        for cell in row:
            if cell.row == 1:
                if not hasattr(model, cell.value):
                    return False
                fields[cell.column] = cell.value
                continue
            field = getattr(model, fields[cell.column])
            if isinstance(getattr(model, fields[cell.column]), ForeignKeyField):
                for_value = DataBaseUtils.get_record(field.rel_model, dict([('uuid', cell.value)]))
                value[fields[cell.column]] = for_value
            else:
                value[fields[cell.column]] = cell.value
        if len(value) > 0:
            data.append(value)
    for value in data:
        DataBaseUtils.get_or_insert(model, value)
    return True


def export_table(model_name):
    model = DataBaseUtils.get_model(model_name)
    model_meta = model._meta
    fields = [key for key in model_meta.fields]
    fields.remove('id')
    data = model.select()
    wb = Workbook()
    ws = wb.active
    ws.title = model_name
    for col, field in enumerate(fields):
        ws.cell(row=1, column=col + 1, value=field)
    row = 2
    for attribute in data:
        for col, field in enumerate(fields):
            val = getattr(attribute, field)
            if isinstance(val, Model):
                val = getattr(val, 'uuid')
            ws.cell(row=row, column=col + 1, value=str(val))
        row += 1
    wb.save(f'expimp/{model_name}.xlsx')
